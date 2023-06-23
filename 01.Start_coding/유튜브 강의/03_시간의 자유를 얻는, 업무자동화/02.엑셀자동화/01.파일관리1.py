import openpyxl

# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
ws = wb.active

# 시트 이름 변경
ws.title = '자동화로만든겅미'

# 새로운 시트 생성
ws = wb.create_sheet('2030.01')             # ws에 넣었으므로 생성한 시트가 활성화됨.

# 모든 시트 이름 출력
print(wb.sheetnames)            


# 시트 삭제
del wb['자동화로만든겅미']

# 엑셀 저장
wb.save(r'startcoding\유튜브 강의\03_시간의 자유를 얻는, 업무자동화\02.엑셀자동화\거래처A매입현황.xlsx')