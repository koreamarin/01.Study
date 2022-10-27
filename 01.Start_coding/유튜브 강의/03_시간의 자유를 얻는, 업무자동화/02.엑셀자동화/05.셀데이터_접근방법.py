import openpyxl

# save_path = 'startcoding/유튜브 강의/03_시간의 자유를 얻는, 업무자동화/02.엑셀자동화/스마트스토어.xlsx'
save_path = r'startcoding\유튜브 강의\03_시간의 자유를 얻는, 업무자동화\02.엑셀자동화\스마트스토어.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True)              # data_only=True 를 넣으면 엑셀 함수가 아니라 데이터형식으로 가져올 수 있다.

# data 시트 선택
ws = wb['data']

# 01. 모든 셀 데이터 가져오기
# -> 행과 열 개수를 아는 경우
# for row in range(1, 9+1) :
#     for column in range(1, 5+1) :
#         print(ws.cell(row, column).value, end=" ")     # ws.cell(row,column).value 는 해당 셀의 값에 접근하는 명령어이다. print(???, end=" ")는 ???를 출력한 후 마지막에 한 칸을 띄우라는 명령어이다.
#     print() # 한 칸 띄어주기용

# -> 행과 열 개수를 모르는 경우             개쩌러!!!
# for row in range(1, ws.max_row + 1) :               # 최대 행 개수까지
#     for column in range(1, ws.max_column + 1) :     # 최대 열 개수까지 실행
#         print(ws.cell(row, column).value, end=" ")     # ws.cell(row,column).value 는 해당 셀의 값에 접근하는 명령어이다. print(???, end=" ")는 ???를 출력한 후 마지막에 한 칸을 띄우라는 명령어이다.
#     print() # 한 칸 띄어주기용

# 모든 행 가져오기
# for row in ws.iter_rows() :             # iter_rows() 괄호안에 아무것도 안쓰면 모든 행을 가져와서 row변수안에 넣어 실행한다.
#     print(row)

# # 모든 열 가져오기
# for column in ws.iter_cols() :
#     print(column)

# # 2번째 행부터 가져오기
# for row in ws.iter_rows(min_row=2) :
#     print(row)

# # 2번째 행부터 5번째 행까지 가져오기.
# for row in ws.iter_rows(min_row=2, max_row=5) :
#     print(row)

# 2-4행, 2-4열 가져와서 출력하기
for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4) :       # for row in ws.iter_rows(2,4,2,4) :
    for cell in row :                     # row 데이터를 cell에 넣어버림
        print(cell, end=" ")              # cell 위치 값을 출력
        print(cell.value, end=" ")        # cell 데이터값을 출력.
    print()

# #자료형.replace(문자열1, 문자열2)은 자료형에 들어있는 문자열1을 문자열2로 대체하여준다.
# present_price = present_price.replace(',', '')