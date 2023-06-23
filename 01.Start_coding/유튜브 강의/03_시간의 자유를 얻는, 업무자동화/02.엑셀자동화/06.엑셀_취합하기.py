# 3개의 엑셀 데이터 취합하기
# 쿠팡, 스스, 11번가 엑셀 데이터 취합

import openpyxl

# 새로운 엑셀 파일 생성
total_wb = openpyxl.Workbook()

# 현재 활성화된 시트 생성
total_ws = total_wb.active

# 시트 이름 변경
total_ws.title = "data"

# 헤더 추가
total_ws.append(['순번', '제품명', '가격','수량','합계'])

# 데이터 파일
file_list = ['11번가','스마트스토어','쿠팡']

for file in file_list:
    wb = openpyxl.load_workbook(f'03_시간의 자유를 얻는, 업무자동화/02.엑셀자동화/{file}.xlsx', data_only=True)     # data_only엑셀 결과값만들 가져오게 하는 것
    ws = wb.active
    for row in ws.iter_rows(min_row=2) :
        data=[]                                 # data 리스트 생성
        for cell in row:                        # row(행)에 있는 값들을 한개씩 불러와 Cell에 넣으면서 반복문 실행
            data.append(cell.value)             # cell에 있는 값을 data 리스트에 한개씩 추가
        total_ws.append(data)                   # data리스트를 total_ws 엑셀시트에 행(row)으로 추가

# # 순번 업데이트 방법 1
# for row in total_ws.iter_rows(min_row=2, max_col=1) : 
#     for cell in row:
#         cell.value = row[0].row - 1

# 순번 업데이트 방법 2
i=0
for cell in total_ws['A'] :
    if i != 0:
            cell.value = i
    i = i + 1


total_wb.save('03_시간의 자유를 얻는, 업무자동화/02.엑셀자동화/total.xlsx')