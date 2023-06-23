import openpyxl

save_fpath = r'startcoding\유튜브 강의\03_시간의 자유를 얻는, 업무자동화\02.엑셀자동화\거래처A매입현황.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_fpath)

# 활성화된 시트 선택
ws = wb.active

# 데이터 추가(1) - 셀명 직집입력하여 데이터 추가 방법
ws['A1'] = '날짜'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'
ws['E1'] = '합계'

# 데이터 추가(2) - 행렬 순서 입력하여 데이터 추가 방법
ws.cell(row=2, column=1, value='2030-01-01')            # (2행, 1열, 데이터 값)
ws.cell(row=2, column=2, value='게이밍 마우스')          # ws.cell(2,2) = '게이밍 마우스' 로도 쓸 수 있다. 완전히 돆같은 형식.
ws.cell(row=2, column=3, value=50000)            
ws.cell(row=2, column=4, value=30)
ws.cell(row=2, column=5, value='=C2*D2')                # 엑셀 함수를 값으로 입력.


# 데이터 추가(3)
ws.append(['2030-01-03','기계식 키보드',120000,15,'=C3*D3'])    # 새로운 행으로 리스트형태의 데이터를 추가한다.

# 데이터 수정
ws['C2'] = 40000
ws['D2'] = 40

# 데이터 삭제
del ws['A3']

# 엑셀 저장
wb.save(save_fpath)

