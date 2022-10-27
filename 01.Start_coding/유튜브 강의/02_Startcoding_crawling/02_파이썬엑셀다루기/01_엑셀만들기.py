import openpyxl

# 1) 엑셀 만들기, 엑셀 파일 생성
wb = openpyxl.Workbook()    # work_book

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')  # work_sheet

# 3) 데이터 추가/수정하기
ws['A1'] = '참가번호'       # 1행 A열에 참가번호 입력
ws['B1'] = '성명'           # 1행 B열에 성명 입력

ws['A2'] = 1
ws['B2'] = '오일남'

# 4) 엑셀 저장하기
# \를 이스케이프문자(\)로 인식시켜주기 위해서는 원래 \\ 이런식으로 이스케이프 문자 두개를 써야한다
# 하지만 문자열 앞에 r을 붙여넣으면 따옴표 안에 있는 문자를 그냥 문자형태로 취급할 수 있도록 할 수 있어서
# 이스케이프 문자를 1개만 쓸 수 있다.
# r 을 쓰지 않았을 때 ('C:\\Users\\USER\\OneDrive\\바탕 화면') or ('C:/Users/USER/OneDrive/바탕 화면') 
wb.save(r'C:\Users\USER\OneDrive\바탕 화면\startcoding\Startcoding_crawling\02_파이썬엑셀다루기\참가자_data.xlsx')


